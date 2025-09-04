import os
import sys
import argparse
from typing import List, Tuple

import asyncio


def normalize_excel_path(path: str) -> str:
    return os.path.abspath(path)


async def update_dummy_columns(
    database_url: str,
    rows: List[Tuple[int, str, int]],
) -> int:
    import asyncpg  # imported lazily to avoid dependency if not needed

    conn = await asyncpg.connect(database_url)
    try:
        # Ensure columns exist
        await conn.execute('ALTER TABLE public."dummy" ADD COLUMN IF NOT EXISTS "주문처명" text;')
        await conn.execute('ALTER TABLE public."dummy" ADD COLUMN IF NOT EXISTS "오더번호" integer;')

        # Batch update
        await conn.executemany(
            'UPDATE public."dummy" SET "주문처명" = $2, "오더번호" = $3, updated_at = NOW() WHERE id = $1;',
            rows,
        )
        return len(rows)
    finally:
        await conn.close()


def load_rows_from_excel(excel_path: str) -> List[Tuple[int, str, int]]:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)

    candidates = []
    for ws in wb.worksheets:
        # Find header row (assume first non-empty row)
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if row and any(cell is not None for cell in row):
                header_row = [str(c).strip() if c is not None else '' for c in row]
                break
        if not header_row:
            continue

        # Required headers
        header_to_idx = {name: i for i, name in enumerate(header_row)}
        required = ['id', '주문처명', '오더번호']
        if not all(k in header_to_idx for k in required):
            continue

        id_idx = header_to_idx['id']
        buyer_idx = header_to_idx['주문처명']
        order_idx = header_to_idx['오더번호']

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            try:
                rid = row[id_idx]
                buyer = row[buyer_idx]
                order_no = row[order_idx]

                if rid is None:
                    continue
                rid_int = int(str(rid).strip())
                buyer_str = None if buyer is None else str(buyer).strip()
                # allow empty buyer; still update order if provided
                if order_no is None or str(order_no).strip() == '':
                    # treat empty as NULL -> skip setting numeric if not provided
                    # we still include in batch with None to set NULL
                    order_int = None
                else:
                    order_int = int(str(order_no).strip())

                candidates.append((rid_int, buyer_str, order_int))
            except Exception:
                # skip malformed row
                continue

        if candidates:
            # Prefer the first worksheet that matches
            break

    return candidates


def main() -> int:
    parser = argparse.ArgumentParser(description='Update public.dummy from Excel (id, 주문처명, 오더번호).')
    parser.add_argument('--excel', required=True, help='Path to Excel file (e.g., masterdb/시연용(인풋)_CBAM 최종.xlsx)')
    parser.add_argument('--database-url', default=os.getenv('DATABASE_URL'), help='PostgreSQL URL')
    args = parser.parse_args()

    if not args.database_url:
        print('ERROR: DATABASE_URL not provided or set in env', file=sys.stderr)
        return 2

    excel_path = normalize_excel_path(args.excel)
    if not os.path.exists(excel_path):
        print(f'ERROR: Excel not found: {excel_path}', file=sys.stderr)
        return 2

    rows = load_rows_from_excel(excel_path)
    if not rows:
        print('ERROR: No updatable rows found (requires headers: id, 주문처명, 오더번호).', file=sys.stderr)
        return 3

    affected = asyncio.run(update_dummy_columns(args.database_url, rows))
    print(f'OK: updated rows = {affected}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())


