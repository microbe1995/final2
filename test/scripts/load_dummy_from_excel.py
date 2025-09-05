#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 파일을 읽어 Railway PostgreSQL의 public.dummy 테이블로 적재하는 스크립트.

엑셀 헤더는 스크린샷과 유사한 한글 컬럼명을 가정하되, 유연한 동의어 매핑을 지원한다.
필수 컬럼: 로트번호, 생산품명, 생산수량, 투입일, 종료일, 공정, 투입물명, 수량
선택 컬럼: 주문처명, 오더번호, 생산수량_단위/생산수량.단위, 투입물_단위/투입물.단위, 단위

사용 예:
  $env:DATABASE_URL='postgresql://...'; 
  python load_dummy_from_excel.py --excel "masterdb/시연용(인풋)_CBAM 최종.xlsx"
"""

from __future__ import annotations

import argparse
import asyncio
import os
import sys
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, date


HEADER_SYNONYMS: Dict[str, Tuple[str, ...]] = {
    '주문처명': ('주문처명', '주문처', '발주처', '고객사', '거래처'),
    '오더번호': ('오더번호', '오더 번호', '로트번호', '로트 번호', 'LOT', 'Lot', 'lot', 'Lot No', 'LOT No', 'LotNo', 'Lot Number', 'lot number'),
    '로트번호': ('로트번호', 'LOT', 'Lot', 'lot', '로트 번호'),
    '생산품명': ('생산품명', '제품명', '생산품'),
    '생산수량': ('생산수량', '생산 수량'),
    '투입일': ('투입일', '시작일', '입력일'),
    '종료일': ('종료일', '완료일', '출고일'),
    '공정': ('공정', '프로세스', 'Process'),
    '투입물명': ('투입물명', '원료/연료명', '원료명', '연료명', '투입물'),
    '수량': ('수량', '투입량', '투입 수량'),
    '투입물_단위': ('투입물_단위', '투입물.단위', '투입 단위', '단위'),
    '단위': ('단위',),
}


def normalize_header_map(headers: List[str]) -> Dict[str, int]:
    h2i = {str(h).strip(): i for i, h in enumerate(headers)}
    result: Dict[str, int] = {}
    for canonical, alts in HEADER_SYNONYMS.items():
        for a in alts:
            if a in h2i:
                result[canonical] = h2i[a]
                break
    return result


def parse_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    s = s.replace('.', '-').replace('/', '-')
    try:
        # allow YYYY-MM-DD or YYYY-MM or YYYY/MM/DD
        parts = s.split('-')
        if len(parts) == 2:
            s = f"{parts[0]}-{parts[1]}-01"
        dt = datetime.fromisoformat(s)
        return dt.date()
    except Exception:
        return None


def to_numeric(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace(',', '')
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def extract_rows_from_excel(excel_path: str) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)
    rows: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        # find header row (first non-empty)
        header = None
        for r in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            if r and any(c is not None for c in r):
                header = [str(c).strip() if c is not None else '' for c in r]
                break
        if not header:
            continue
        hmap = normalize_header_map(header)
        # minimally require 핵심 헤더들
        required = ('로트번호', '생산품명', '생산수량', '투입일', '종료일', '공정', '투입물명', '수량')
        if not all(k in hmap for k in required):
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or all(c is None for c in r):
                continue
            get = lambda key: (r[hmap[key]] if key in hmap and hmap[key] < len(r) else None)
            item: Dict[str, Any] = {
                '주문처명': get('주문처명'),
                '오더번호': get('오더번호'),
                '로트번호': get('로트번호'),
                '생산품명': get('생산품명'),
                '생산수량': to_numeric(get('생산수량')),
                '투입일': parse_date(get('투입일')),
                '종료일': parse_date(get('종료일')),
                '공정': get('공정'),
                '투입물명': get('투입물명'),
                '수량': to_numeric(get('수량')),
                '투입물_단위': get('투입물_단위') or 'ton',
                '단위': get('투입물_단위') or get('단위') or 'ton',
            }
            rows.append(item)
        if rows:
            break
    return rows


async def insert_rows(database_url: str, items: List[Dict[str, Any]]) -> int:
    import asyncpg
    conn = await asyncpg.connect(database_url)
    try:
        # batch insert
        q = (
            'INSERT INTO public."dummy" '
            '("주문처명","오더번호","로트번호","생산품명","생산수량","생산수량_단위","투입일","종료일","공정","투입물명","수량","투입물_단위","단위") '
            'VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13)'
        )
        def S(v: Any) -> Optional[str]:
            if v is None:
                return None
            s = str(v).strip()
            return s if s != '' else None
        def I(v: Any) -> Optional[int]:
            if v is None:
                return None
            if isinstance(v, int):
                return v
            s = str(v).strip()
            if s.isdigit():
                return int(s)
            digits = ''.join(ch for ch in s if ch.isdigit())
            return int(digits) if digits else None
        values = [
            (
                S(i.get('주문처명')),
                I(i.get('오더번호')),
                S(i.get('로트번호')),
                S(i.get('생산품명')),
                i.get('생산수량'),
                S(i.get('생산수량_단위')),
                i.get('투입일'),
                i.get('종료일'),
                S(i.get('공정')),
                S(i.get('투입물명')),
                i.get('수량'),
                S(i.get('투입물_단위')),
                S(i.get('단위')),
            )
            for i in items
        ]
        await conn.executemany(q, values)
        return len(values)
    finally:
        await conn.close()


def main(argv: List[str]) -> int:
    parser = argparse.ArgumentParser(description='Load Excel rows into public.dummy table.')
    parser.add_argument('--excel', default=os.path.abspath(os.path.join('masterdb', '시연용(인풋)_CBAM 최종.xlsx')))
    parser.add_argument('--database-url', default=os.getenv('DATABASE_URL'))
    args = parser.parse_args(argv)

    if not os.path.exists(args.excel):
        print(f'ERROR: Excel not found: {args.excel}', file=sys.stderr)
        return 2
    if not args.database_url:
        print('ERROR: DATABASE_URL not provided or set in env', file=sys.stderr)
        return 2

    items = extract_rows_from_excel(args.excel)
    if not items:
        print('ERROR: No rows detected from excel (check header names).', file=sys.stderr)
        return 3

    inserted = asyncio.run(insert_rows(args.database_url, items))
    print(f'OK: inserted rows = {inserted}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main(sys.argv[1:]))


