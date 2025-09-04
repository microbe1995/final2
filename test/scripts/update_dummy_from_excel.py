import os
import sys
import argparse
from typing import List, Tuple

import asyncio


def normalize_excel_path(path: str) -> str:
    return os.path.abspath(path)


async def update_dummy_columns(
    database_url: str,
    rows: List[Tuple[int, str, int]],
) -> int:
    import asyncpg  # imported lazily to avoid dependency if not needed

    conn = await asyncpg.connect(database_url)
    try:
        # Ensure columns exist
        await conn.execute('ALTER TABLE public."dummy" ADD COLUMN IF NOT EXISTS "주문처명" text;')
        await conn.execute('ALTER TABLE public."dummy" ADD COLUMN IF NOT EXISTS "오더번호" integer;')

        # Batch update
        await conn.executemany(
            'UPDATE public."dummy" SET "주문처명" = $2, "오더번호" = $3, updated_at = NOW() WHERE id = $1;',
            rows,
        )
        return len(rows)
    finally:
        await conn.close()


def load_rows_from_excel(excel_path: str) -> List[Tuple[int, str, int]]:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)

    candidates = []
    for ws in wb.worksheets:
        # Find header row robustly: look up to 30 rows for a row containing likely header tokens
        header_row = None
        def _norm(s):
            s = str(s) if s is not None else ''
            return s.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ').strip()
        header_candidates = []
        for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
            if not row or all(c is None for c in row):
                continue
            norm = [_norm(c) for c in row]
            header_candidates.append(norm)
            tokens = set(norm)
            likely = {'id', 'ID', '주문처명', '주문처', '발주처', '오더번호', '로트번호', '로트 번호', 'LOT', 'Lot', 'lot'}
            if tokens.intersection(likely):
                header_row = norm
                break
        if not header_row and header_candidates:
            # fallback: first non-empty
            header_row = header_candidates[0]

        # Build header map with synonyms
        header_to_idx = {name: i for i, name in enumerate(header_row)}
        def _find_idx(names):
            for n in names:
                if n in header_to_idx:
                    return header_to_idx[n]
            return None

        id_idx = _find_idx(['id', 'ID'])
        buyer_idx = _find_idx(['주문처명', '주문처', '발주처', '고객사', '거래처', '주문처 명'])
        order_idx = _find_idx(['오더번호', '오더 번호', '로트번호', '로트 번호', 'LOT', 'Lot', 'lot', 'Lot No', 'LOT No', 'LotNo', 'Lot Number', 'lot number'])

        # Require id and at least one of buyer/order
        if id_idx is None or (buyer_idx is None and order_idx is None):
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            try:
                rid = row[id_idx]
                buyer = row[buyer_idx] if buyer_idx is not None else None
                order_no = row[order_idx] if order_idx is not None else None

                if rid is None:
                    continue
                rid_int = int(str(rid).strip())
                buyer_str = None if buyer is None else str(buyer).strip()
                # parse order number flexibly (allow strings with spaces/newlines)
                if order_no is None or str(order_no).strip() == '':
                    # treat empty as NULL -> skip setting numeric if not provided
                    # we still include in batch with None to set NULL
                    order_int = None
                else:
                    s = str(order_no).strip()
                    # if it's purely digits, cast; otherwise try to extract digits
                    if s.isdigit():
                        order_int = int(s)
                    else:
                        digits = ''.join(ch for ch in s if ch.isdigit())
                        order_int = int(digits) if digits else None

                candidates.append((rid_int, buyer_str, order_int))
            except Exception:
                # skip malformed row
                continue

        if candidates:
            # Prefer the first worksheet that matches
            break

    return candidates


def main() -> int:
    parser = argparse.ArgumentParser(description='Update public.dummy from Excel (id, 주문처명/주문처, 오더번호/로트번호).')
    parser.add_argument('--excel', required=True, help='Path to Excel file (e.g., masterdb/시연용(인풋)_CBAM 최종.xlsx)')
    parser.add_argument('--database-url', default=os.getenv('DATABASE_URL'), help='PostgreSQL URL')
    args = parser.parse_args()

    if not args.database_url:
        print('ERROR: DATABASE_URL not provided or set in env', file=sys.stderr)
        return 2

    excel_path = normalize_excel_path(args.excel)
    if not os.path.exists(excel_path):
        print(f'ERROR: Excel not found: {excel_path}', file=sys.stderr)
        return 2

    rows = load_rows_from_excel(excel_path)
    if not rows:
        print('ERROR: No updatable rows found (need id and one of 주문처명/주문처 or 오더번호/로트번호).', file=sys.stderr)
        return 3

    affected = asyncio.run(update_dummy_columns(args.database_url, rows))
    print(f'OK: updated rows = {affected}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())


